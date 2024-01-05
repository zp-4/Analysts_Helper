import csv
import re
import openpyxl
import xlrd
from openpyxl import Workbook
import argparse
import os
import sys

# Constants for file types and default output filename
CSV = 'csv'
XLSX = 'xlsx'
XLS = 'xls'
DEFAULT_CSV_NAME = 'output.csv'  # Default output file name if none is provided

def strip_quotes(arg):
    """
    Strip quotes from the given command line argument.

    Args:
        arg (str): The command line argument.

    Returns:
        str: The argument with surrounding single or double quotes removed, if present.
    """
    if arg.startswith(("'", '"')) and arg.endswith(("'", '"')):
        return arg[1:-1]
    return arg

class UniversalHandler:
    """
    A universal file handler class to determine and return the appropriate file handler
    based on the file extension.

    Attributes:
        filename (str): The name of the file.
        handler (CSVHandler or ExcelHandler): The determined file handler.
    """
    def __init__(self, filename):
        self.filename = filename
        self.handler = self.get_handler()

    def get_handler(self):
        """
        Determine the handler based on file extension.

        Returns:
            A file handler (CSVHandler or ExcelHandler).

        Raises:
            ValueError: If the file type is unsupported.
        """
        if self.filename.lower().endswith('.csv'):
            return CSVHandler(self.filename)
        elif self.filename.lower().endswith('.xlsx'):
            return ExcelHandler(self.filename, XLSX)
        elif self.filename.lower().endswith('.xls'):
            return ExcelHandler(self.filename, XLS)
        else:
            raise ValueError("Unsupported file type")

class CSVHandler:
    """
    A handler class for CSV files to perform extract, search, and column search operations.

    Attributes:
        filename (str): The name of the file.
    """
    def __init__(self, filename):
        self.filename = filename

    def extract_columns(self, columns, new_filename):
        """
        Extract specified columns from a CSV file and writes them to a new file.

        Args:
            columns (list of str): Columns to extract.
            new_filename (str): The filename for the new file with extracted columns.

        Returns:
            str: Success or error message.
        """
        try:
            with open(self.filename, mode='r', newline='', encoding='utf-8') as file, open(new_filename, mode='w', newline='', encoding='utf-8') as new_file:
                reader = csv.DictReader(file)
                writer = csv.DictWriter(new_file, fieldnames=columns)
                writer.writeheader()
                for row in reader:
                    writer.writerow({column: row[column] for column in columns})
            return f"Columns {columns} extracted to {new_filename} successfully."
        except Exception as e:
            return f"Error occurred: {e}"

    def search_csv(self, pattern):
        """
        Searches the entire CSV for a given string or regex and returns the rows that match.

        Args:
            pattern (str): String or regex pattern to search for.

        Returns:
            list: List of rows that match the pattern.
        """
        matches = []
        try:
            with open(self.filename, mode='r', newline='', encoding='utf-8') as file:
                reader = csv.reader(file)
                for row in reader:
                    if any(re.search(pattern, cell) for cell in row):
                        matches.append(row)
            return matches
        except Exception as e:
            return f"Error occurred: {e}"

    def search_column(self, column, pattern, output_choice='print', new_filename=DEFAULT_CSV_NAME):
        """
        Searches a given column for a string or regex and returns unique results.
        Can print in terminal or create a new CSV.

        Args:
            column (str): Column to search.
            pattern (str): String or regex pattern to search for.
            output_choice (str): 'print' to display in terminal, 'csv' to create a new file.
            new_filename (str): The filename for the new file with search results.

        Returns:
            set or str: Unique search results or a success/error message.
        """
        unique_results = set()
        try:
            with open(self.filename, mode='r', newline='', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    if re.search(pattern, row[column]):
                        unique_results.add(row[column])

            if output_choice == 'print':
                for result in unique_results:
                    print(result)
            elif output_choice == 'csv':
                output_path = os.path.join(os.getcwd(), new_filename)  # Using provided name or default for file paths
                with open(output_path, mode='w', newline='', encoding='utf-8') as new_file:
                    writer = csv.writer(new_file)
                    for result in unique_results:
                        writer.writerow([result])
                return f"Search results saved to {output_path}"
            return unique_results
        except Exception as e:
            return f"Error occurred: {e}"

class ExcelHandler:
    """
    A handler class for Excel (.xlsx and .xls) files to perform extract, search, and column search operations.

    Attributes:
        filename (str): The name of the file.
        file_type (str): The type of Excel file ('xlsx' or 'xls').
    """
    def __init__(self, filename, file_type):
        self.filename = filename
        self.file_type = file_type

    def extract_columns(self, columns, new_filename):
        """
        Extract specified columns from an Excel file and writes them to a new file.

        Args:
            columns (list of str): Columns to extract.
            new_filename (str): The filename for the new file with extracted columns.

        Returns:
            str: Success or error message.
        """
        try:
            if self.file_type == XLSX:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                new_wb = Workbook()
                new_ws = new_wb.active
                for row in ws.iter_rows(values_only=True):
                    new_ws.append([row[columns.index(col)] for col in columns if col in columns])
                new_wb.save(new_filename)
            elif self.file_type == XLS:
                workbook = xlrd.open_workbook(self.filename)
                sheet = workbook.sheet_by_index(0)
                new_wb = Workbook()
                new_ws = new_wb.active
                for rx in range(sheet.nrows):
                    row = sheet.row(rx)
                    new_ws.append([row[columns.index(col)].value for col in columns if col in columns])
                new_wb.save(new_filename)
            return f"Columns {columns} extracted to {new_filename} successfully."
        except Exception as e:
            return f"Error occurred: {e}"

    def search_csv(self, pattern):
        """
        Searches the entire Excel file for a given string or regex and returns the rows that match.

        Args:
            pattern (str): String or regex pattern to search for.

        Returns:
            list: List of rows that match the pattern.
        """
        matches = []
        try:
            if self.file_type == XLSX:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if any(re.search(pattern, str(cell)) for cell in row):
                        matches.append(row)
            elif self.file_type == XLS:
                workbook = xlrd.open_workbook(self.filename)
                sheet = workbook.sheet_by_index(0)
                for rx in range(sheet.nrows):
                    row = sheet.row(rx)
                    if any(re.search(pattern, str(cell.value)) for cell in row):
                        matches.append([cell.value for cell in row])
            return matches
        except Exception as e:
            return f"Error occurred: {e}"

    def search_column(self, column, pattern, output_choice='print', new_filename=DEFAULT_CSV_NAME):
        """
        Searches a given column in an Excel file for a string or regex and returns unique results.
        Can print in terminal or create a new CSV.

        Args:
            column (str): Column to search.
            pattern (str): String or regex pattern to search for.
            output_choice (str): 'print' to display in terminal, 'csv' to create a new file.
            new_filename (str): The filename for the new file with search results.

        Returns:
            set or str: Unique search results or a success/error message.
        """
        unique_results = set()
        try:
            if self.file_type == XLSX:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                col_idx = openpyxl.utils.cell.column_index_from_string(column) - 1
                for row in ws.iter_rows(values_only=True):
                    cell_value = str(row[col_idx])
                    if re.search(pattern, cell_value):
                        unique_results.add(cell_value)
            elif self.file_type == XLS:
                workbook = xlrd.open_workbook(self.filename)
                sheet = workbook.sheet_by_index(0)
                col_idx = None
                for idx, col_val in enumerate(sheet.row_values(0)):
                    if col_val == column:
                        col_idx = idx
                        break
                if col_idx is not None:
                    for rx in range(1, sheet.nrows):
                        cell_value = str(sheet.cell_value(rx, col_idx))
                        if re.search(pattern, cell_value):
                            unique_results.add(cell_value)

            if output_choice == 'print':
                for result in unique_results:
                    print(result)
            elif output_choice == 'csv':
                output_path = os.path.join(os.getcwd(), new_filename)  # Using provided name or default for file paths
                with open(output_path, mode='w', newline='', encoding='utf-8') as new_file:
                    writer = csv.writer(new_file)
                    for result in unique_results:
                        writer.writerow([result])
                return f"Search results saved to {output_path}"
            return unique_results
        except Exception as e:
            return f"Error occurred: {e}"

def parse_arguments():
    """
    Parse command-line arguments using argparse library.

    Returns:
        argparse.Namespace: Parsed arguments.
    """
    parser = argparse.ArgumentParser(description="Handle CSV and Excel files for various operations.")
    parser.add_argument('filename', help='The file to process')
    parser.add_argument('--extract', nargs='+', help='Columns to extract. Provide column names separated by spaces.')
    parser.add_argument('--newfile', help='Name of the new file when extracting columns or saving search results')
    parser.add_argument('--search', help='String or regex to search in the entire file')
    parser.add_argument('--searchcol', help='Column to perform search on')
    parser.add_argument('--pattern', help='String or regex pattern to search for')
    parser.add_argument('--output', choices=['print', 'csv'], default='print', help='Output choice for search results')
    return parser.parse_args()

def main():
    """
    The main function to execute the file handler operations based on command-line arguments.
    """
    args = parse_arguments()

    # Strip quotes from arguments as necessary
    filename = strip_quotes(args.filename)
    searchcol = strip_quotes(args.searchcol) if args.searchcol else None
    pattern = strip_quotes(args.pattern) if args.pattern else None
    output = strip_quotes(args.output)
    newfile = strip_quotes(args.newfile) if args.newfile else DEFAULT_CSV_NAME

    # Instantiate the appropriate handler based on file type
    handler = UniversalHandler(filename).handler

    # Perform operations based on command-line arguments
    if args.extract and newfile:
        print(handler.extract_columns(args.extract, newfile))
    elif args.search:
        print(handler.search_csv(pattern))
    elif searchcol and pattern:
        print(handler.search_column(searchcol, pattern, output, new_filename=newfile))

if __name__ == "__main__":
    main()
