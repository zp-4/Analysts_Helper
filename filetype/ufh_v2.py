import csv
import re
import openpyxl
import xlrd
from openpyxl import Workbook
from odf.opendocument import load
from odf.table import Table, TableRow, TableCell
import argparse
import os

# Constants for file types and default output filename
CSV = 'csv'
XLSX = 'xlsx'
XLS = 'xls'
ODS = 'ods'
DEFAULT_CSV_NAME = 'output.csv'

def strip_quotes(arg):
    if arg.startswith(("'", '"')) and arg.endswith(("'", '"')):
        return arg[1:-1]
    return arg

class UniversalHandler:
    def __init__(self, filename):
        self.filename = filename
        self.handler = self.get_handler()

    def get_handler(self):
        if self.filename.lower().endswith('.csv'):
            return CSVHandler(self.filename)
        elif self.filename.lower().endswith('.xlsx'):
            return ExcelHandler(self.filename, XLSX)
        elif self.filename.lower().endswith('.xls'):
            return ExcelHandler(self.filename, XLS)
        elif self.filename.lower().endswith('.ods'):
            return ODSHandler(self.filename)
        else:
            raise ValueError("Unsupported file type")

class CSVHandler:
    def __init__(self, filename):
        self.filename = filename

    def extract_columns(self, columns, new_filename):
        try:
            with open(self.filename, mode='r', newline='', encoding='utf-8') as file, open(new_filename, mode='w', newline='', encoding='utf-8') as new_file:
                reader = csv.DictReader(file)
                writer = csv.DictWriter(new_file, fieldnames=columns)
                writer.writeheader()
                for row in reader:
                    writer.writerow({column: row[column] for column in columns})
            return f"Columns {columns} extracted to {new_filename} successfully."
        except Exception as e:
            return f"Error occurred: {e}"

    def search_csv(self, pattern):
        matches = []
        try:
            with open(self.filename, mode='r', newline='', encoding='utf-8') as file:
                reader = csv.reader(file)
                for row in reader:
                    if any(re.search(pattern, cell) for cell in row):
                        matches.append(row)
            return matches
        except Exception as e:
            return f"Error occurred: {e}"

    def search_column(self, column, pattern, output_choice='print', new_filename=DEFAULT_CSV_NAME):
        unique_results = set()
        try:
            with open(self.filename, mode='r', newline='', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    if re.search(pattern, row[column]):
                        unique_results.add(row[column])

            if output_choice == 'print':
                for result in unique_results:
                    print(result)
            elif output_choice == 'csv':
                output_path = os.path.join(os.getcwd(), new_filename)
                with open(output_path, mode='w', newline='', encoding='utf-8') as new_file:
                    writer = csv.writer(new_file)
                    for result in unique_results:
                        writer.writerow([result])
                return f"Search results saved to {output_path}"
            return unique_results
        except Exception as e:
            return f"Error occurred: {e}"

class ExcelHandler:
    def __init__(self, filename, file_type):
        self.filename = filename
        self.file_type = file_type

    def extract_columns(self, columns, new_filename):
        try:
            if self.file_type == XLSX:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                new_wb = Workbook()
                new_ws = new_wb.active
                for row in ws.iter_rows(values_only=True):
                    new_ws.append([row[col] for col in columns if col in columns])
                new_wb.save(new_filename)
            elif self.file_type == XLS:
                workbook = xlrd.open_workbook(self.filename)
                sheet = workbook.sheet_by_index(0)
                new_wb = Workbook()
                new_ws = new_wb.active
                for rx in range(sheet.nrows):
                    row = sheet.row(rx)
                    new_ws.append([row[col].value for col in columns if col in columns])
                new_wb.save(new_filename)
            return f"Columns {columns} extracted to {new_filename} successfully."
        except Exception as e:
            return f"Error occurred: {e}"

    def search_csv(self, pattern):
        matches = []
        try:
            if self.file_type == XLSX:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if any(re.search(pattern, str(cell)) for cell in row):
                        matches.append(row)
            elif self.file_type == XLS:
                workbook = xlrd.open_workbook(self.filename)
                sheet = workbook.sheet_by_index(0)
                for rx in range(sheet.nrows):
                    row = sheet.row(rx)
                    if any(re.search(pattern, str(cell.value)) for cell in row):
                        matches.append([cell.value for cell in row])
            return matches
        except Exception as e:
            return f"Error occurred: {e}"

    def search_column(self, column, pattern, output_choice='print', new_filename=DEFAULT_CSV_NAME):
        unique_results = set()
        try:
            if self.file_type == XLSX:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    cell_value = str(row[column])
                    if re.search(pattern, cell_value):
                        unique_results.add(cell_value)
            elif self.file_type == XLS:
                workbook = xlrd.open_workbook(self.filename)
                sheet = workbook.sheet_by_index(0)
                for rx in range(1, sheet.nrows):
                    cell_value = str(sheet.cell(rx, column).value)
                    if re.search(pattern, cell_value):
                        unique_results.add(cell_value)

            if output_choice == 'print':
                for result in unique_results:
                    print(result)
            elif output_choice == 'csv':
                output_path = os.path.join(os.getcwd(), new_filename)
                with open(output_path, mode='w', newline='', encoding='utf-8') as new_file:
                    writer = csv.writer(new_file)
                    for result in unique_results:
                        writer.writerow([result])
                return f"Search results saved to {output_path}"
            return unique_results
        except Exception as e:
            return f"Error occurred: {e}"

class ODSHandler:
    def __init__(self, filename):
        self.filename = filename

    def extract_columns(self, columns, new_filename):
        try:
            ods_file = load(self.filename)
            sheet = ods_file.spreadsheet.getElementsByType(Table)[0]
            new_ods = load()  # Create a new blank ODS file
            new_sheet = Table()  # Create a new sheet
            for row in sheet.getElementsByType(TableRow):
                new_row = TableRow()  # Create a new row for the new sheet
                for i, cell in enumerate(row.getElementsByType(TableCell)):
                    if i in columns:
                        new_row.addElement(cell)
                new_sheet.addElement(new_row)
            new_ods.spreadsheet.addElement(new_sheet)
            new_ods.save(new_filename)
            return f"Columns {columns} extracted to {new_filename} successfully."
        except Exception as e:
            return f"Error occurred: {e}"

    def search_ods(self, pattern):
        matches = []
        try:
            ods_file = load(self.filename)
            sheet = ods_file.spreadsheet.getElementsByType(Table)[0]
            for row in sheet.getElementsByType(TableRow):
                matched_row = []
                for cell in row.getElementsByType(TableCell):
                    cell_value = str(cell)
                    if re.search(pattern, cell_value):
                        matched_row.append(cell_value)
                if matched_row:
                    matches.append(matched_row)
            return matches
        except Exception as e:
            return f"Error occurred: {e}"

    def search_column(self, column, pattern, output_choice='print', new_filename=DEFAULT_CSV_NAME):
        unique_results = set()
        try:
            ods_file = load(self.filename)
            sheet = ods_file.spreadsheet.getElementsByType(Table)[0]
            for row in sheet.getElementsByType(TableRow):
                cells = row.getElementsByType(TableCell)
                if len(cells) > column:
                    cell_value = str(cells[column])
                    if re.search(pattern, cell_value):
                        unique_results.add(cell_value)

            if output_choice == 'print':
                for result in unique_results:
                    print(result)
            elif output_choice == 'csv':
                output_path = os.path.join(os.getcwd(), new_filename)
                with open(output_path, mode='w', newline='', encoding='utf-8') as new_file:
                    writer = csv.writer(new_file)
                    for result in unique_results:
                        writer.writerow([result])
                return f"Search results saved to {output_path}"
            return unique_results
        except Exception as e:
            return f"Error occurred: {e}"

def parse_arguments():
    parser = argparse.ArgumentParser(description="Handle CSV, Excel, and ODS files for various operations.")
    parser.add_argument('filename', help='The file to process')
    parser.add_argument('--extract', nargs='+', help='Columns to extract. Provide column names separated by spaces.')
    parser.add_argument('--newfile', help='Name of the new file when extracting columns or saving search results')
    parser.add_argument('--search', help='String or regex to search in the entire file')
    parser.add_argument('--searchcol', help='Column to perform search on')
    parser.add_argument('--pattern', help='String or regex pattern to search for')
    parser.add_argument('--output', choices=['print', 'csv'], default='print', help='Output choice for search results')
    return parser.parse_args()

def main():
    args = parse_arguments()

    filename = strip_quotes(args.filename)
    searchcol = strip_quotes(args.searchcol) if args.searchcol else None
    pattern = strip_quotes(args.pattern) if args.pattern else None
    output = strip_quotes(args.output)
    newfile = strip_quotes(args.newfile) if args.newfile else DEFAULT_CSV_NAME

    handler = UniversalHandler(filename).handler

    if args.extract and newfile:
        print(handler.extract_columns(args.extract, newfile))
    elif args.search:
        print(handler.search_ods(pattern))
    elif searchcol and pattern:
        print(handler.search_column(searchcol, pattern, output, new_filename=newfile))

if __name__ == "__main__":
    main()
