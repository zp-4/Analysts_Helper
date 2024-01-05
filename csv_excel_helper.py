import csv
import re
import openpyxl
import xlrd
from openpyxl import Workbook
import argparse


# Constants for file types
CSV = 'csv'
XLSX = 'xlsx'
XLS = 'xls'

class UniversalHandler:
    def __init__(self, filename):
        self.filename = filename
        self.handler = self.get_handler()

    def get_handler(self):
        if self.filename.endswith('.csv'):
            return CSVHandler(self.filename)
        elif self.filename.endswith('.xlsx'):
            return ExcelHandler(self.filename, XLSX)
        elif self.filename.endswith('.xls'):
            return ExcelHandler(self.filename, XLS)
        else:
            raise ValueError("Unsupported file type")

class CSVHandler:
    def __init__(self, filename):
        self.filename = filename

    def extract_columns(self, columns, new_filename):
        try:
            with open(self.filename, mode='r') as file, open(new_filename, mode='w', newline='') as new_file:
                reader = csv.DictReader(file)
                writer = csv.DictWriter(new_file, fieldnames=columns)
                writer.writeheader()
                for row in reader:
                    writer.writerow({column: row[column] for column in columns})
            return f"Columns {columns} extracted to {new_filename} successfully."
        except Exception as e:
            return f"Error occurred: {e}"

    def search_csv(self, pattern):
        matches = []
        try:
            with open(self.filename, mode='r') as file:
                reader = csv.reader(file)
                for row in reader:
                    if any(re.search(pattern, cell) for cell in row):
                        matches.append(row)
            return matches
        except Exception as e:
            return f"Error occurred: {e}"

    def search_column(self, column, pattern, output_choice='print'):
        unique_results = set()
        try:
            with open(self.filename, mode='r') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    if re.search(pattern, row[column]):
                        unique_results.add(row[column])

            if output_choice == 'print':
                for result in unique_results:
                    print(result)
            elif output_choice == 'csv':
                with open(f'{column}_search_results.csv', mode='w', newline='') as new_file:
                    writer = csv.writer(new_file)
                    for result in unique_results:
                        writer.writerow([result])
                return f"Search results saved to {column}_search_results.csv"
            return unique_results
        except Exception as e:
            return f"Error occurred: {e}"

class ExcelHandler:
    def __init__(self, filename, file_type):
        self.filename = filename
        self.file_type = file_type

    def extract_columns(self, columns, new_filename):
        try:
            if self.file_type == XLSX:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                new_wb = Workbook()
                new_ws = new_wb.active
                for row in ws.iter_rows(values_only=True):
                    new_ws.append([row[columns.index(col)] for col in columns if col in columns])
                new_wb.save(new_filename)
            elif self.file_type == XLS:
                workbook = xlrd.open_workbook(self.filename)
                sheet = workbook.sheet_by_index(0)
                new_wb = Workbook()
                new_ws = new_wb.active
                for rx in range(sheet.nrows):
                    row = sheet.row(rx)
                    new_ws.append([row[columns.index(col)].value for col in columns if col in columns])
                new_wb.save(new_filename)
            return f"Columns {columns} extracted to {new_filename} successfully."
        except Exception as e:
            return f"Error occurred: {e}"

    def search_csv(self, pattern):
        matches = []
        try:
            if self.file_type == XLSX:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if any(re.search(pattern, str(cell)) for cell in row):
                        matches.append(row)
            elif self.file_type == XLS:
                workbook = xlrd.open_workbook(self.filename)
                sheet = workbook.sheet_by_index(0)
                for rx in range(sheet.nrows):
                    row = sheet.row(rx)
                    if any(re.search(pattern, str(cell.value)) for cell in row):
                        matches.append([cell.value for cell in row])
            return matches
        except Exception as e:
            return f"Error occurred: {e}"

    def search_column(self, column, pattern, output_choice='print'):
        unique_results = set()
        try:
            if self.file_type == XLSX:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                col_idx = openpyxl.utils.cell.column_index_from_string(column) - 1
                for row in ws.iter_rows(values_only=True):
                    cell_value = str(row[col_idx])
                    if re.search(pattern, cell_value):
                        unique_results.add(cell_value)
            elif self.file_type == XLS:
                workbook = xlrd.open_workbook(self.filename)
                sheet = workbook.sheet_by_index(0)
                col_idx = None
                for idx, col_val in enumerate(sheet.row_values(0)):
                    if col_val == column:
                        col_idx = idx
                        break
                if col_idx is not None:
                    for rx in range(1, sheet.nrows):
                        cell_value = str(sheet.cell_value(rx, col_idx))
                        if re.search(pattern, cell_value):
                            unique_results.add(cell_value)

            if output_choice == 'print':
                for result in unique_results:
                    print(result)
            elif output_choice == 'csv':
                with open(f'{column}_search_results.csv', mode='w', newline='') as new_file:
                    writer = csv.writer(new_file)
                    for result in unique_results:
                        writer.writerow([result])
                return f"Search results saved to {column}_search_results.csv"
            return unique_results
        except Exception as e:
            return f"Error occurred: {e}"



# Define the command-line arguments
def parse_arguments():
    parser = argparse.ArgumentParser(description="Handle CSV and Excel files for various operations.")
    parser.add_argument('filename', help='The file to process')
    parser.add_argument('--extract', nargs='+', help='Columns to extract. Provide column names separated by spaces.')
    parser.add_argument('--newfile', help='Name of the new file when extracting columns or saving search results')
    parser.add_argument('--search', help='String or regex to search in the entire file')
    parser.add_argument('--searchcol', help='Column to perform search on')
    parser.add_argument('--pattern', help='String or regex pattern to search for')
    parser.add_argument('--output', choices=['print', 'csv'], default='print', help='Output choice for search results')
    return parser.parse_args()

def main():
    args = parse_arguments()

    # Instantiate the appropriate handler based on file type
    handler = UniversalHandler(args.filename).handler

    # Perform operations based on command-line arguments
    if args.extract and args.newfile:
        print(handler.extract_columns(args.extract, args.newfile))
    elif args.search:
        print(handler.search_csv(args.search))
    elif args.searchcol and args.pattern:
        print(handler.search_column(args.searchcol, args.pattern, args.output))

if __name__ == "__main__":
    main()
